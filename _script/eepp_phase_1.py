"""
EEPP Phase 1
------------
Pulls ERCOT solar, wind, system forecast, HROC, and MORA data, merges/derives
metrics, saves styled Excel + PNG plots, and (optionally) posts to Slack.
"""
from __future__ import annotations
import logging, os, shutil, sys
from datetime import datetime
from io import BytesIO
import matplotlib.image as mpimg
import matplotlib.pyplot as plt
import pandas as pd
import requests, zipfile
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from tqdm import tqdm
from webdriver_manager.chrome import ChromeDriverManager

LOG_FILE = "eepp_ph_1.log"
PRODUCTION_DIR = "production"
ARCHIVE_DIR = os.path.join(PRODUCTION_DIR, "archive")
ICONS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "icons")
DEFAULT_LOGO = os.path.normpath(os.path.join(ICONS_DIR, "logo.png"))
HEADLESS = True
URL_SOLAR = "https://www.ercot.com/mp/data-products/data-product-details?id=NP4-737-CD"
URL_WIND = "https://www.ercot.com/mp/data-products/data-product-details?id=NP4-732-CD"
URL_SYSFCST = "https://www.ercot.com/mp/data-products/data-product-details?id=NP3-560-CD"
URL_HROC = "https://www.ercot.com/mp/data-products/data-product-details?id=NP3-233-CD"
URL_MORA = "https://www.ercot.com/gridinfo/resource"
SLACK_TOKEN = os.environ.get("SLACK_TOKEN")
SLACK_CHANNEL = "#eepp"

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s",
                    filename=LOG_FILE, filemode="a")
log = logging.getLogger("eepp_phase_1")
log.info("****** BEGIN - EEPP Phase 1 script.")

def ensure_dirs():
    os.makedirs(PRODUCTION_DIR, exist_ok=True)
    os.makedirs(ARCHIVE_DIR, exist_ok=True)

def is_file_older_than_today(p:str)->bool:
    from datetime import datetime as dt
    return dt.fromtimestamp(os.path.getctime(p)).date() < dt.now().date()

def archive_old_files():
    for fn in os.listdir(PRODUCTION_DIR):
        src = os.path.join(PRODUCTION_DIR, fn); dst = os.path.join(ARCHIVE_DIR, fn)
        if os.path.isfile(src) and is_file_older_than_today(src):
            if os.path.exists(dst): os.remove(dst)
            shutil.move(src, dst)
    log.info("Moved old files to archive.")

def get_download_link(driver, url:str, link_text="zip", file_type="zip")->str:
    log.info("Navigating to URL: %s", url); driver.get(url)
    try:
        if file_type == "xlsx":
            xpath = ("//a[contains(@href, '.xlsx') and "
                     "contains(@title, 'Monthly Outlook for Resource Adequacy (MORA)')]")
            link = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, xpath)))
        else:
            link = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.LINK_TEXT, link_text)))
        href = link.get_attribute("href"); log.info("Found download link: %s", href); return href
    except Exception as exc:
        log.exception("Error finding download link: %s", exc); raise

def download_and_extract_zip(url:str, extract_to:str=".")->str:
    log.info("Downloading ZIP from %s", url)
    r = requests.get(url, timeout=60); r.raise_for_status()
    with zipfile.ZipFile(BytesIO(r.content)) as zf:
        zf.extractall(extract_to); names = zf.namelist()
    return names[0]

def process_file(driver, page_url:str, new_filename:str, progress:tqdm, link_text="zip", file_type="zip", download_dir=".")->str:
    href = get_download_link(driver, page_url, link_text, file_type)
    if not href.startswith("http"): href = "https://www.ercot.com"+href
    out_path = os.path.join(download_dir, new_filename)
    if file_type=="zip":
        extracted = download_and_extract_zip(href, extract_to=download_dir)
        os.replace(os.path.join(download_dir, extracted), out_path)
    else:
        r = requests.get(href, timeout=60); r.raise_for_status()
        with open(out_path, "wb") as f: f.write(r.content)
    progress.update(1); log.info("Processed %s", new_filename); return out_path

def send_slack_notification(message:str, files=None):
    if not SLACK_TOKEN: log.info("SLACK_TOKEN not set; skipping Slack."); return
    client = WebClient(token=SLACK_TOKEN)
    try:
        resp = client.chat_postMessage(channel=SLACK_CHANNEL, text=message, mrkdwn=True)
        ts = resp["ts"]
        for path in files or []:
            with open(path, "rb") as fh:
                client.files_upload_v2(channel=resp["channel"], thread_ts=ts, file=fh, title=os.path.basename(path))
        log.info("Message sent to Slack: %s", resp["channel"])
    except SlackApiError as e:
        log.error("Slack error: %s", e.response.get("error", "unknown"))

def plot_combined_forecast(df: pd.DataFrame, logo_path:str)->str:
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(20, 20))
    ax1.plot(df["datetime"], df["Forecasted Demand"], label="Forecasted Demand")
    ax1.plot(df["datetime"], df["Forecasted Wind Supply"], label="Forecasted Wind Supply")
    ax1.plot(df["datetime"], df["Forecasted Solar Supply"], label="Forecasted Solar Supply")
    ax1.plot(df["datetime"], df["Forecasted Renewables Output"], label="Forecasted Renewables Output")
    ax1.plot(df["datetime"], df["Dispatchable Supply"], label="Dispatchable Supply")
    ax1.set_xlabel("Date"); ax1.set_ylabel("MW")
    ax1.set_title("ERCOT 7-Day Supply/Demand Forecast", fontsize=16)
    ax1.legend(loc="upper left"); ax1.grid(True)
    bar_width = 1/24
    ax2.bar(df["datetime"], df["Forecasted Thermal Reserve Margin"], color="lightgray", alpha=0.85, width=bar_width,
            label="Forecasted Thermal Reserve Margin (Bar)", zorder=1)
    ax3 = ax2.twinx()
    ax3.plot(df["datetime"], df["Forecasted Net Load"], label="Forecasted Net Load")
    ax3.plot(df["datetime"], df["Forecasted Net Thermal Capacity"], label="Forecasted Net Thermal Capacity")
    ax2.set_xlabel("Date"); ax2.set_ylabel("MW (Thermal Reserve Margin)")
    ax3.set_ylabel("MW (Net Load / Net Thermal Capacity)")
    ax2.set_title("ERCOT 7-Day Forecasted Thermal Reserve Margin", fontsize=16)
    b1,l1 = ax2.get_legend_handles_labels(); l2,lbl2 = ax3.get_legend_handles_labels()
    leg = ax2.legend(b1 + l2, l1 + lbl2, loc="upper left", fontsize=10); leg.set_zorder(5); ax2.grid(True)
    fig.text(0.95, 0.04, datetime.now().isoformat(), ha="right", fontsize=12)
    try:
        if os.path.exists(logo_path):
            logo = mpimg.imread(logo_path); fig.figimage(logo, 50, 50, zorder=10, alpha=0.5)
    except Exception: pass
    out_png = os.path.join(PRODUCTION_DIR, f"ercot_combined_forecast_{datetime.now().strftime('%Y-%m-%d_T%H_%M_%S')}.png")
    plt.savefig(out_png, bbox_inches="tight"); plt.close(fig); return out_png

def main():
    ensure_dirs(); archive_old_files()
    ts = datetime.now().strftime("%Y-%m-%d_T%H_%M_%S")
    solar_fn, wind_fn = f"solar_{ts}.csv", f"wind_{ts}.csv"
    sys_fn, hroc_fn = f"sysfcst_{ts}.csv", f"hroc_{ts}.csv"
    mora_fn = f"mora_{ts}.xlsx"
    total = 10; progress = tqdm(total=total, desc="Phase 1: Processing Files")
    service = Service(ChromeDriverManager().install())
    opts = webdriver.ChromeOptions(); opts.add_argument("--headless=new")
    driver = webdriver.Chrome(service=service, options=opts)
    try:
        solar_path = process_file(driver, URL_SOLAR, solar_fn, progress, download_dir=PRODUCTION_DIR)
        wind_path = process_file(driver, URL_WIND, wind_fn, progress, download_dir=PRODUCTION_DIR)
        sys_path  = process_file(driver, URL_SYSFCST, sys_fn, progress, download_dir=PRODUCTION_DIR)
        hroc_path = process_file(driver, URL_HROC, hroc_fn, progress, download_dir=PRODUCTION_DIR)
        mora_path = process_file(driver, URL_MORA, mora_fn, progress, link_text="xlsx", file_type="xlsx",
                                 download_dir=PRODUCTION_DIR)
    finally:
        driver.quit()
    df_solar = pd.read_csv(solar_path, parse_dates=["DELIVERY_DATE"]); progress.update(1)
    df_wind  = pd.read_csv(wind_path,  parse_dates=["DELIVERY_DATE"]); progress.update(1)
    df_sys   = pd.read_csv(sys_path,   parse_dates=["DeliveryDate"]);   progress.update(1)
    df_hroc  = pd.read_csv(hroc_path,  parse_dates=["Date"]);           progress.update(1)
    df_mora  = pd.read_excel(mora_path, sheet_name="Capacity by Resource Category"); progress.update(1)
    eactor_dhhrsr_value = df_mora.iloc[2, 3]
    wind_solar = df_solar.merge(df_wind, on=["DELIVERY_DATE","HOUR_ENDING"])
    df_sys["HOUR_ENDING"] = df_sys.HourEnding.astype(str).str.split(":").str[0].astype(int)
    wind_solar_hroc = wind_solar.merge(df_hroc, left_on=["DELIVERY_DATE","HOUR_ENDING"],
                                       right_on=["Date","HourEnding"])
    merged = df_sys.merge(wind_solar_hroc, left_on=["DeliveryDate","HOUR_ENDING"],
                          right_on=["DELIVERY_DATE","HOUR_ENDING"])
    merged.rename(columns={
        "SystemTotal":"Forecasted Demand",
        "COP_HSL_SYSTEM_WIDE_y":"Forecasted Wind Supply",
        "COP_HSL_SYSTEM_WIDE_x":"Forecasted Solar Supply",
    }, inplace=True)
    merged["Forecasted Renewables Output"] = merged["Forecasted Wind Supply"] + merged["Forecasted Solar Supply"]
    merged["Dispatchable Supply"] = merged["Forecasted Demand"] - merged["Forecasted Renewables Output"]
    merged["EACTOR_DHHRSR"] = eactor_dhhrsr_value
    merged["total_resource_outages"] = (
        merged["TotalResourceMWZoneSouth"] + merged["TotalResourceMWZoneNorth"] +
        merged["TotalResourceMWZoneWest"]  + merged["TotalResourceMWZoneHouston"]
    )
    merged["Forecasted Net Load"] = merged["Forecasted Demand"] - merged["Forecasted Renewables Output"]
    merged["Forecasted Net Thermal Capacity"] = merged["EACTOR_DHHRSR"] - merged["total_resource_outages"]
    merged["Forecasted Thermal Reserve Margin"] = merged["Forecasted Net Thermal Capacity"] - merged["Forecasted Net Load"]
    merged["datetime"] = pd.to_datetime(merged.DeliveryDate.astype(str) + " " + (merged.HOUR_ENDING - 1).astype(str) + ":00")
    png_path = plot_combined_forecast(merged, DEFAULT_LOGO); progress.update(1)
    xlsx_path = os.path.join(PRODUCTION_DIR, f"fcst_wind_solar_hroc_mora_{datetime.now().strftime('%Y-%m-%d_T%H_%M_%S')}.xlsx")
    merged.to_excel(xlsx_path, index=False)
    wb = load_workbook(xlsx_path); ws = wb.active
    header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for c in ws[1]: c.fill = header_fill; c.font = header_font
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(10, min(max_len + 2, 60))
    wb.save(xlsx_path)
    from os.path import basename
    send_slack_notification(f"*7-Day Forecasts*\\nNew ERCOT forecasts processed.\\n\\n{basename(png_path)}\\n{basename(xlsx_path)}",
                            files=[png_path])
    progress.update(1); progress.close(); log.info("END - EEPP Phase 1 script. ******")

if __name__ == "__main__":
    main()
